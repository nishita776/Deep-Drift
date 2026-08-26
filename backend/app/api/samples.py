import os
import io
import csv
import uuid
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app import models, schemas
from app.database import get_db, SessionLocal
from app.config import settings
from app.conservation import conservation_status_for
from app.pipeline.runner import run_pipeline

router = APIRouter(prefix="/samples", tags=["samples"])


@router.post("", response_model=schemas.SampleOut)
def upload_sample(
    name: str = Form(...),
    sample_type: str = Form("field"),
    marker_gene: str = Form("18S"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload a FASTQ file + metadata. Returns the created sample record
    with its id — the frontend uses that id for every subsequent call."""
    if marker_gene.upper() not in ("18S", "COI"):
        raise HTTPException(422, "marker_gene must be '18S' or 'COI'")

    os.makedirs(settings.storage_dir, exist_ok=True)
    file_path = os.path.join(settings.storage_dir, f"{uuid.uuid4()}_{file.filename}")
    with open(file_path, "wb") as f:
        f.write(file.file.read())

    sample = models.Sample(
        name=name, sample_type=sample_type, marker_gene=marker_gene,
        upload_path=file_path, status="uploaded",
    )
    db.add(sample)
    db.commit()
    db.refresh(sample)
    return sample


def _run_pipeline_background(sample_id: str, job_id: str):
    """Runs in a background thread with its OWN db session — never share
    a session across threads/requests."""
    db = SessionLocal()
    try:
        run_pipeline(db, sample_id, job_id)
    finally:
        db.close()


@router.post("/{sample_id}/run", response_model=schemas.JobOut)
def run_sample(sample_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    sample = db.get(models.Sample, sample_id)
    if not sample:
        raise HTTPException(404, "Sample not found")

    job = models.Job(sample_id=sample_id, status="pending")
    db.add(job)
    db.commit()
    db.refresh(job)

    background_tasks.add_task(_run_pipeline_background, sample_id, job.id)
    return job


@router.get("/{sample_id}/results", response_model=schemas.SampleResultsOut)
def get_results(sample_id: str, db: Session = Depends(get_db)):
    sample = db.get(models.Sample, sample_id)
    if not sample:
        raise HTTPException(404, "Sample not found")

    matches = (
        db.query(models.TaxaMatch)
        .join(models.ASV)
        .filter(models.ASV.sample_id == sample_id, models.TaxaMatch.status == "confident_match")
        .all()
    )
    known_taxa = [
        schemas.TaxaMatchOut(
            asv_id=m.asv_id, sequence_preview=m.asv.sequence[:40] + "...",
            count=m.asv.count, status=m.status, matched_taxon=m.matched_taxon,
            identity_score=m.identity_score, database_source=m.database_source,
            conservation_status=conservation_status_for(m.matched_taxon),
        )
        for m in matches
    ]
    total_reads = sum(a.count for a in sample.asvs)

    return schemas.SampleResultsOut(sample=sample, known_taxa=known_taxa, total_reads=total_reads)


@router.get("/{sample_id}/novel-clusters", response_model=list[schemas.ClusterOut])
def get_novel_clusters(sample_id: str, db: Session = Depends(get_db)):
    sample = db.get(models.Sample, sample_id)
    if not sample:
        raise HTTPException(404, "Sample not found")
    clusters = (
        db.query(models.Cluster)
        .filter(models.Cluster.sample_id == sample_id)
        .order_by(models.Cluster.novelty_score.desc())
        .all()
    )
    return clusters


@router.get("/{sample_id}/biodiversity", response_model=schemas.BiodiversityOut)
def get_biodiversity(sample_id: str, db: Session = Depends(get_db)):
    sample = db.get(models.Sample, sample_id)
    if not sample:
        raise HTTPException(404, "Sample not found")
    if not sample.biodiversity:
        raise HTTPException(409, "Biodiversity metrics not yet computed — has the pipeline finished running?")
    return sample.biodiversity


@router.get("/compare")
def compare_samples(ids: str, db: Session = Depends(get_db)):
    """ids is a comma-separated list, e.g. /samples/compare?ids=abc,def,ghi"""
    sample_ids = [s.strip() for s in ids.split(",") if s.strip()]
    results = {}
    for sid in sample_ids:
        sample = db.get(models.Sample, sid)
        if not sample:
            continue
        results[sid] = {
            "name": sample.name,
            "status": sample.status,
            "biodiversity": schemas.BiodiversityOut.model_validate(sample.biodiversity) if sample.biodiversity else None,
        }
    return results


def _gather_export_data(db: Session, sample_id: str):
    """Shared by both CSV and Excel export so the two formats can never
    drift out of sync with each other. Returns (sample, known_taxa_rows,
    novel_cluster_rows, biodiversity) or raises HTTPException(404)."""
    sample = db.get(models.Sample, sample_id)
    if not sample:
        raise HTTPException(404, "Sample not found")

    matches = (
        db.query(models.TaxaMatch)
        .join(models.ASV)
        .filter(models.ASV.sample_id == sample_id, models.TaxaMatch.status == "confident_match")
        .all()
    )
    known_taxa_rows = [
        {
            "asv_id": m.asv_id,
            "matched_taxon": m.matched_taxon or "",
            "identity_score": m.identity_score if m.identity_score is not None else "",
            "database_source": m.database_source or "",
            "read_count": m.asv.count,
        }
        for m in matches
    ]

    clusters = (
        db.query(models.Cluster)
        .filter(models.Cluster.sample_id == sample_id)
        .order_by(models.Cluster.novelty_score.desc())
        .all()
    )
    novel_cluster_rows = [
        {
            "placeholder_id": c.placeholder_id,
            "rank_prediction": c.rank_prediction or "",
            "nearest_reference": c.nearest_reference or "",
            "novelty_score": c.novelty_score,
            "member_count": c.member_count,
            "total_reads": c.total_reads,
        }
        for c in clusters
    ]

    return sample, known_taxa_rows, novel_cluster_rows, sample.biodiversity


@router.get(
    "/{sample_id}/export",
    summary="Download a CSV or Excel report (known taxa + novel clusters + biodiversity)",
)
def export_sample(sample_id: str, format: str = "csv", db: Session = Depends(get_db)):
    """Downloadable report combining known taxa, novel clusters, and
    biodiversity metrics. `format` is 'csv' (default) or 'xlsx'.

    CSV: one file, sections separated by a blank line + header row per
    section (simplest format that opens correctly everywhere, including
    Excel/Google Sheets/plain text editors).
    XLSX: three separate sheets — Known Taxa / Novel Clusters / Biodiversity.
    """
    if format not in ("csv", "xlsx"):
        raise HTTPException(422, "format must be 'csv' or 'xlsx'")

    sample, known_taxa_rows, novel_cluster_rows, biodiversity = _gather_export_data(db, sample_id)
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in sample.name) or sample_id
    filename_base = f"{safe_name}_report"

    if format == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        writer.writerow(["Sample", sample.name])
        writer.writerow(["Sample ID", sample.id])
        writer.writerow(["Marker Gene", sample.marker_gene])
        writer.writerow(["Status", sample.status])
        writer.writerow([])

        writer.writerow(["KNOWN TAXA"])
        writer.writerow(["ASV ID", "Matched Taxon", "Identity Score", "Database Source", "Read Count"])
        for r in known_taxa_rows:
            writer.writerow([r["asv_id"], r["matched_taxon"], r["identity_score"],
                              r["database_source"], r["read_count"]])
        writer.writerow([])

        writer.writerow(["CANDIDATE NOVEL CLUSTERS"])
        writer.writerow(["Cluster ID", "Rank Prediction", "Nearest Reference",
                          "Novelty Score", "Member Count", "Total Reads"])
        for r in novel_cluster_rows:
            writer.writerow([r["placeholder_id"], r["rank_prediction"], r["nearest_reference"],
                              r["novelty_score"], r["member_count"], r["total_reads"]])
        writer.writerow([])

        writer.writerow(["BIODIVERSITY METRICS"])
        if biodiversity:
            writer.writerow(["Shannon Index", biodiversity.shannon])
            writer.writerow(["Simpson Index", biodiversity.simpson])
            writer.writerow(["Richness", biodiversity.richness])
        else:
            writer.writerow(["Not yet computed"])

        buffer.seek(0)
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    else:  # xlsx
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(
                500,
                "openpyxl is not installed. Run: pip install openpyxl "
                "(it's in requirements.txt — reinstall dependencies).",
            )

        wb = Workbook()

        ws_taxa = wb.active
        ws_taxa.title = "Known Taxa"
        ws_taxa.append(["ASV ID", "Matched Taxon", "Identity Score", "Database Source", "Read Count"])
        for r in known_taxa_rows:
            ws_taxa.append([r["asv_id"], r["matched_taxon"], r["identity_score"],
                             r["database_source"], r["read_count"]])

        ws_clusters = wb.create_sheet("Novel Clusters")
        ws_clusters.append(["Cluster ID", "Rank Prediction", "Nearest Reference",
                             "Novelty Score", "Member Count", "Total Reads"])
        for r in novel_cluster_rows:
            ws_clusters.append([r["placeholder_id"], r["rank_prediction"], r["nearest_reference"],
                                 r["novelty_score"], r["member_count"], r["total_reads"]])

        ws_bio = wb.create_sheet("Biodiversity")
        ws_bio.append(["Sample", sample.name])
        ws_bio.append(["Marker Gene", sample.marker_gene])
        ws_bio.append([])
        if biodiversity:
            ws_bio.append(["Shannon Index", biodiversity.shannon])
            ws_bio.append(["Simpson Index", biodiversity.simpson])
            ws_bio.append(["Richness", biodiversity.richness])
            ws_bio.append([])
            ws_bio.append(["Rarefaction depths", *biodiversity.rarefaction_curve.get("depths", [])])
            ws_bio.append(["Rarefaction richness", *biodiversity.rarefaction_curve.get("richness", [])])
        else:
            ws_bio.append(["Biodiversity metrics not yet computed"])

        # Autosize columns roughly, for a report that looks decent on open
        for ws in (ws_taxa, ws_clusters, ws_bio):
            for col_cells in ws.columns:
                length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )
